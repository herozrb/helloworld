from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
import traceback
import json
from datetime import datetime
import requests
import sseclient
from tqdm import tqdm  # 导入 tqdm 库
import base64  # 添加这行导入
import openpyxl
import time
import threading
import logging
from logging.handlers import RotatingFileHandler

# 在文件开头，获取项目根目录的绝对路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, 
    template_folder=os.path.join(BASE_DIR, 'templates'))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB限制

# 确保上传目录和导出目录都存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'exports'), exist_ok=True)

# 允许的文件类型
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

# 在文件开头添加 API_URL 常量
API_URL = 'https://api.siliconflow.cn/v1/chat/completions'

# 在 API_URL 常量定义后添加日志配置
def setup_logger():
    """配置日志系统"""
    # 创建logs目录
    log_dir = os.path.join(BASE_DIR, 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    # 配置日志文件
    log_file = os.path.join(log_dir, 'app.log')
    
    # 创建logger对象
    logger = logging.getLogger('AIcoding')
    logger.setLevel(logging.DEBUG)
    
    # 创建文件处理器，设置最大文件大小为10MB，最多保留5个备份
    file_handler = RotatingFileHandler(
        log_file,
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5,
        encoding='utf-8'
    )
    
    # 创建控制台处理器
    console_handler = logging.StreamHandler()
    
    # 设置日志格式
    formatter = logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(module)s: %(message)s',
        datefmt='%y/%m/%d/%H/%M/%S'
    )
    
    # 设置处理器的格式
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # 添加处理器到logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

# 创建logger实例
logger = setup_logger()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

def generate_basic_auth_key(username):
    """生成Basic认证的密钥"""
    auth_str = f"{username}:"  # 不需要密码，但保留冒号
    auth_bytes = auth_str.encode('utf-8')
    encoded_auth = base64.b64encode(auth_bytes)
    return f"Basic {encoded_auth.decode('utf-8')}"

def call_model(text, prompt):
    """调用模型进行分析"""
    try:
        headers = {
            'Authorization': f"Bearer sk-usoulvohbnxowonlipznpyefpcafbcitbqvymonrderhirjh",
            'Content-Type': 'application/json'
        }
        
        data = {
            "model": "/Qwen2.5-32B-InstructQwen",
            "messages": [
                {
                    "role": "system",
                    "content": "你是一个专业的文本分析助手，需要帮助分析文本并给出明确的判断结果。请只返回'是'或'否'。"
                },
                {
                    "role": "user",
                    "content": f"请分析以下文本，{prompt}\n\n文本内容：{text}\n\n请只回答'是'或'否'。"
                }
            ],
            "stream": False,
            "max_tokens": 512,
            "stop": ["null"],
            "temperature": 0.7,
            "top_p": 0.7,
            "top_k": 50,
            "frequency_penalty": 0.5,
            "n": 1,
            "response_format": {"type": "text"}
        }
        
        response = requests.post(
            'https://api.siliconflow.cn/v1/chat/completions',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            result = response.json()
            print(f"API Response: {result}")  # 调试信息
            if 'choices' in result and result['choices']:
                message = result['choices'][0].get('message', {})
                if message and 'content' in message:
                    return message['content'].strip()
            print(f"Invalid response format: {result}")
            return None
        else:
            print(f"模型调用错误: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"模型调用错误: {str(e)}")
        return None

def apply_model_analysis(df, prompt):
    """使用模型分析文本"""
    total_rows = len(df)
    matched_count = 0
    sample_records = []
    
    # 使用 tqdm 创建进度条
    with tqdm(total=total_rows, desc="处理进度", unit="行") as pbar:
        for text in df['context_around_keyword'].astype(str):
            result = call_model(text, prompt)
            if result == '是':
                matched_count += 1
                if len(sample_records) < 10:  # 只保存前10个匹配的记录
                    sample_records.append(text)
            pbar.update(1)  # 更新进度条
    
    percentage = (matched_count / total_rows * 100) if total_rows > 0 else 0
    
    return {
        'total_rows': total_rows,
        'matched_count': matched_count,
        'percentage': round(percentage, 2),
        'sample_records': sample_records
    }

def process_data_with_model(text, rules):
    """使用模型处理数据"""
    try:
        # 记录开始时间
        start_time = datetime.now()
        logger.info(f"开始处理数据...")

        # 构建提示词
        prompt = f"""请分析以下对话内容，根据规则进行标注。

规则：
{rules}

对话内容：
{text}

请仅输出标签，多个标签用逗号分隔。"""

        headers = {
            'Authorization': f"Bearer sk-usoulvohbnxowonlipznpyefpcafbcitbqvymonrderhirjh",
            'Content-Type': 'application/json'
        }
        
        data = {
            "model": "deepseek-ai/DeepSeek-R1-Distill-Qwen-32B",
            "messages": [
                {
                    "role": "system",
                    "content": "你是一个专业的对话分析助手，擅长根据规则对对话内容进行标注。"
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "stream": False,
            "max_tokens": 2048,
            "temperature": 0.3,
            "top_p": 0.8,
            "frequency_penalty": 0.3,
            "presence_penalty": 0.1,
            "stop": None,
            "n": 1
        }

        logger.info("正在调用模型...")
        response = requests.post(API_URL, headers=headers, json=data, timeout=60)
        
        # 记录原始响应
        logger.debug(f"模型原始响应: {response.text}")
        
        if response.status_code == 200:
            result = response.json()
            if 'choices' in result and len(result['choices']) > 0:
                content = result['choices'][0]['message']['content']
                logger.info(f"模型返回结果: {content}")
                return content
            else:
                logger.error(f"响应格式无效: {result}")
                return None
        else:
            logger.error(f"请求失败: {response.status_code}")
            return None

    except Exception as e:
        logger.error(f"处理错误: {str(e)}", exc_info=True)
        return None

def process_dataframe(df, rules):
    """处理整个数据框"""
    try:
        # 获取所有列名
        columns = df.columns.tolist()
        
        # 为每列创建新的处理后的列
        processed_df = pd.DataFrame()  # 创建新的数据框
        
        for col in columns:
            # 复制原始列
            processed_df[col] = df[col]
            new_col_name = f"{col}_processed"
            processed_df[new_col_name] = None
            
            # 使用 tqdm 显示处理进度
            with tqdm(total=len(df), desc=f"处理 {col} 列", unit="行") as pbar:
                for idx, value in df[col].astype(str).items():
                    processed_value = process_data_with_model(value, rules)
                    if processed_value is not None:  # 只在处理成功时更新值
                        processed_df.at[idx, new_col_name] = processed_value
                    pbar.update(1)
        
        # 生成输出文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
        os.makedirs(export_dir, exist_ok=True)  # 确保导出目录存在
        output_file = os.path.join(export_dir, f'processed_{timestamp}.xlsx')
        
        # 保存处理后的数据
        processed_df.to_excel(output_file, index=False)
        
        return {
            'success': True,
            'message': '数据处理完成',
            'file': os.path.basename(output_file)
        }
    except Exception as e:
        print(f"数据处理错误: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

def clean_dialogue_data(json_data):
    """清洗对话数据，兼容两种JSON格式"""
    try:
        # 检查是否为空值
        if pd.isna(json_data):
            return {
                'success': True,
                'data': ''  # 返回空字符串
            }
            
        # 预处理JSON字符串
        json_str = str(json_data).strip()
        if json_str.lower() == 'nan':  # 再次检查nan字符串
            return {
                'success': True,
                'data': ''
            }
            
        if json_str.startswith('【') and json_str.endswith('】'):
            json_str = json_str[1:-1]  # 移除中文方括号
        
        # 处理可能的特殊字符
        json_str = json_str.replace('\n', '').replace('\r', '').strip()
        
        # 如果是空字符串，直接返回
        if not json_str:
            return {
                'success': True,
                'data': ''
            }
        
        # 解析JSON数据
        try:
            dialogue_data = json.loads(json_str)
        except json.JSONDecodeError:
            # 如果解析失败，尝试eval转换
            try:
                dialogue_data = eval(json_str)
                if not isinstance(dialogue_data, list):
                    raise ValueError("数据格式不是对话列表")
            except Exception as e:
                print(f"数据格式转换失败: {str(e)}")
                print(f"原始数据: {json_str[:200]}...")
                raise
        
        # 如果对话数据为空，返回空字符串
        if not dialogue_data:
            return {
                'success': True,
                'data': ''
            }
        
        # 构建对话文本和标签列表
        dialogue_text = []
        all_labels = []  # 存储所有标签
        
        for entry in dialogue_data:
            try:
                # 获取说话角色（兼容两种格式）
                if 'roleName' in entry:
                    speaker = '客户' if entry['roleName'] == 'customer' else '坐席'
                else:
                    speaker = '客户' if entry['role'] == 'customer' else '坐席'
                
                # 获取对话内容
                content = entry['text'].strip()
                
                # 获取时间（兼容两种格式）
                if isinstance(entry.get('startTime'), (int, float)) and isinstance(entry.get('endTime'), (int, float)):
                    time = f"{entry['startTime']}-{entry['endTime']}秒"
                else:
                    # 处理可能的字符串格式
                    start_time = float(str(entry['startTime']).replace('秒', ''))
                    end_time = float(str(entry['endTime']).replace('秒', ''))
                    time = f"{start_time}-{end_time}秒"
                
                # 获取标签信息（兼容两种格式）
                labels = []
                # 新版格式
                if 'sentenceParamMap' in entry and 'labelList' in entry['sentenceParamMap']:
                    labels = entry['sentenceParamMap']['labelList']
                # 旧版格式
                elif 'semanticViolationItems' in entry:
                    for item in entry['semanticViolationItems']:
                        if item.get('callScriptName') != '无明显标签':
                            labels.append(item['callScriptName'])
                    if entry.get('semanticViolationDetection') != '#未理解#':
                        labels.append(entry['semanticViolationDetection'])
                
                # 构建对话文本（不包含标签信息）
                dialogue_line = f"{speaker}({time}): {content}"
                dialogue_text.append(dialogue_line)
                
                # 收集标签
                if labels:
                    all_labels.extend(labels)
            
            except Exception as e:
                print(f"处理单条对话时出错: {str(e)}")
                print(f"问题数据: {entry}")
                continue
        
        if not dialogue_text:
            raise ValueError("没有成功解析任何对话内容")
        
        # 将所有对话内容用换行符连接
        full_dialogue = '\n'.join(dialogue_text)
        # 将所有标签去重并用逗号连接
        unique_labels = list(set(all_labels))
        labels_text = ', '.join(unique_labels) if unique_labels else ''
        
        return {
            'success': True,
            'data': full_dialogue,
            'labels': labels_text
        }
        
    except Exception as e:
        print(f"处理错误: {str(e)}")
        return {
            'success': False,
            'error': f"数据处理错误: {str(e)}\n请确保数据格式正确，并且包含完整的对话信息。"
        }

def process_long_dialogue(df, rules):
    """处理长对话并使用大模型打标"""
    try:
        # 复制原始DataFrame
        result_df = df.copy()
        # 添加新列用于存储模型标注结果
        result_df['label_MOXING'] = None
        
        # 筛选出通话时长超过50秒的对话
        long_dialogues = result_df[result_df['talking_seconds'] > 50]
        
        if long_dialogues.empty:
            print("没有找到超过50秒的对话")
            return {
                'success': True,
                'message': '处理完成，没有找到需要处理的长对话',
                'df': result_df
            }
        
        print(f"找到 {len(long_dialogues)} 条长对话需要处理")
        
        # 使用tqdm显示处理进度
        with tqdm(total=len(long_dialogues), desc="处理长对话", unit="条") as pbar:
            for idx in long_dialogues.index:
                try:
                    # 获取对话内容
                    content = result_df.at[idx, 'content_new']
                    if pd.isna(content) or not content.strip():
                        continue
                    
                    # 调用大模型处理
                    processed_value = process_data_with_model(content, rules)
                    if processed_value:
                        result_df.at[idx, 'label_MOXING'] = processed_value
                    
                except Exception as e:
                    print(f"处理第 {idx} 行数据时出错: {str(e)}")
                    continue
                finally:
                    pbar.update(1)
        
        return {
            'success': True,
            'message': f'成功处理 {len(long_dialogues)} 条长对话',
            'df': result_df
        }
        
    except Exception as e:
        print(f"处理长对话时出错: {str(e)}")
        return {
            'success': False,
            'error': str(e)
        }

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        print("开始处理上传请求")
        if 'file' not in request.files:
            return jsonify({'error': '没有选择文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400
        
        if file and allowed_file(file.filename):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            print(f"文件已保存到: {filepath}")
            
            try:
                if filepath.endswith('.csv'):
                    df = pd.read_csv(filepath, encoding='utf-8')
                else:
                    df = pd.read_excel(filepath)
                
                if df.empty:
                    return jsonify({'error': '文件内容为空'}), 400
                
                analysis_type = request.form.get('analysisType', 'keyword')
                print(f"分析类型: {analysis_type}")
                
                # 获取规则输入
                rules = request.form.get('rules', '').strip()
                
                # 根据不同的分析类型进行处理
                if analysis_type == 'clean':
                    print("开始数据清洗")
                    # 复制原始DataFrame
                    result_df = df.copy()
                    # 添加新列
                    result_df['content_new'] = None
                    result_df['labels_new'] = None  # 添加标签列
                    
                    # 处理每一行数据
                    for index, row in df.iterrows():
                        print(f"处理第 {index + 1} 行数据")
                        try:
                            # 获取content列的数据
                            content = row.get('content', '')
                            if pd.isna(content):
                                continue
                                
                            result = clean_dialogue_data(content)
                            if result['success']:
                                if result['data']:
                                    result_df.at[index, 'content_new'] = result['data']
                                if result['labels']:
                                    result_df.at[index, 'labels_new'] = result['labels']
                        
                        except Exception as e:
                            print(f"处理第 {index + 1} 行数据时出错: {str(e)}")
                            continue
                    
                    # 生成输出文件
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
                    os.makedirs(export_dir, exist_ok=True)
                    output_file = os.path.join(export_dir, f'cleaned_dialogue_{timestamp}.xlsx')
                    
                    print(f"保存清洗后的数据到: {output_file}")
                    # 设置Excel单元格格式以保持换行
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        result_df.to_excel(writer, index=False)
                        worksheet = writer.sheets['Sheet1']
                        
                        # 设置两个新列的格式
                        for col_name in ['content_new', 'labels_new']:
                            col_idx = result_df.columns.get_loc(col_name) + 1
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            # 对话内容列宽度设置为100，标签列宽度设置为50
                            worksheet.column_dimensions[col_letter].width = 100 if col_name == 'content_new' else 50
                            for row in worksheet.iter_rows(min_row=2, max_row=len(result_df) + 1, min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    
                    return jsonify({
                        'success': True,
                        'message': '数据清洗完成',
                        'download_url': f"/download/{os.path.basename(output_file)}"
                    })
                
                elif analysis_type == 'process':
                    # 检查必要的列是否存在
                    required_columns = ['content_new', 'talking_seconds']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        return jsonify({'error': f'文件中缺少以下列: {", ".join(missing_columns)}'}), 400
                    
                    if not rules:
                        return jsonify({'error': '请输入标注规则'}), 400
                    
                    # 处理长对话
                    result = process_long_dialogue(df, rules)
                    if not result['success']:
                        return jsonify({'error': result['error']}), 400
                    
                    # 生成输出文件
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
                    os.makedirs(export_dir, exist_ok=True)
                    output_file = os.path.join(export_dir, f'processed_dialogue_{timestamp}.xlsx')
                    
                    # 保存结果
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        result['df'].to_excel(writer, index=False)
                        worksheet = writer.sheets['Sheet1']
                        
                        # 设置新列的格式
                        col_idx = result['df'].columns.get_loc('label_MOXING') + 1
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        worksheet.column_dimensions[col_letter].width = 50
                        for row in worksheet.iter_rows(min_row=2, max_row=len(result['df']) + 1, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    
                    return jsonify({
                        'success': True,
                        'message': result['message'],
                        'download_url': f"/download/{os.path.basename(output_file)}"
                    })
                
                else:  # keyword 或 model 分析
                    if not rules:
                        return jsonify({'error': '请输入规则'}), 400
                    
                    # 检查必要的列
                    content_column = 'context_around_keyword'
                    if content_column not in df.columns:
                        return jsonify({'error': f'文件中缺少"{content_column}"列'}), 400
                    
                    if analysis_type == 'keyword':
                        stats = apply_rules(df, rules)
                    else:  # model
                        stats = apply_model_analysis(df, rules)
                    
                    return jsonify({
                        'success': True,
                        'stats': stats
                    })
            
            except Exception as e:
                print(f"处理文件时出错: {str(e)}")
                return jsonify({'error': f'文件处理错误: {str(e)}'}), 400
            
        return jsonify({'error': '不支持的文件类型'}), 400
    
    except Exception as e:
        print(f"上传过程出错: {str(e)}")
        return jsonify({'error': f'上传过程错误: {str(e)}'}), 500

def apply_rules(df, keyword):
    """使用关键词匹配分析"""
    try:
        content_column = 'context_around_keyword'
        
        # 清理文本内容
        df['cleaned_content'] = df[content_column].astype(str)
        
        # 总行数
        total_rows = len(df)
        
        # 找到包含关键词的行
        matched_rows = df[df['cleaned_content'].str.contains(keyword, na=False)]
        matched_count = len(matched_rows)
        
        # 计算占比
        percentage = (matched_count / total_rows * 100) if total_rows > 0 else 0
        
        # 获取前10个匹配的记录
        sample_records = matched_rows['cleaned_content'].head(10).tolist()
        
        return {
            'total_rows': total_rows,
            'matched_count': matched_count,
            'percentage': round(percentage, 2),
            'sample_records': sample_records
        }
    except Exception as e:
        print(f"规则应用错误: {str(e)}")
        raise

@app.route('/export', methods=['POST'])
def export_results():
    try:
        data = request.json
        if not data or 'stats' not in data:
            return jsonify({'error': '没有数据可导出'}), 400
        
        # 创建导出目录
        export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
        os.makedirs(export_dir, exist_ok=True)
        
        # 生成导出文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_file = os.path.join(export_dir, f'分析结果_{timestamp}.xlsx')
        
        # 创建数据框
        stats = data['stats']
        summary_df = pd.DataFrame([{
            '总数据量': stats['total_rows'],
            '匹配数量': stats['matched_count'],
            '匹配占比': f"{stats['percentage']}%",
            '搜索关键词': data['keyword']
        }])
        
        records_df = pd.DataFrame(stats['sample_records'], columns=['匹配记录'])
        
        # 使用ExcelWriter保存到不同的sheet
        with pd.ExcelWriter(export_file) as writer:
            summary_df.to_excel(writer, sheet_name='统计摘要', index=False)
            records_df.to_excel(writer, sheet_name='匹配记录', index=True)
        
        return jsonify({
            'success': True,
            'file': os.path.basename(export_file)
        })
    except Exception as e:
        print(f"导出错误: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    try:
        export_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
        file_path = os.path.join(export_dir, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'error': f'文件不存在: {filename}'}), 404
            
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': f'下载失败: {str(e)}'}), 500

if __name__ == '__main__':
    # 移除守护进程相关代码，直接运行 Flask 应用
    app.run(debug=True, port=5001)